"""
Модуль: modified_csv_exporter
Назначение: Экспорт обновлённого CSV/XLSX файла с подсветкой старых и новых UID.

Функционал:
- Сохраняет иерархию с UID в XLSX (с цветами) или CSV (с пометками)
- Использует настройки из config.json:
  - Формат: xlsx/csv
  - Суффикс файла
  - Колонки (через config["csv"]["headers"])
- Подсвечивает зелёным UID, которые уже были в исходном CSV
- Оставляет чёрным цветом новые (сгенерированные) UID
- Добавляет столбец 'status' в CSV при необходимости

Используется в process_file после генерации XML.
"""

import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Set, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None


def _read_original_uids(csv_path: Path, config: Dict, logger: Optional[logging.Logger] = None) -> Set[str]:
    """
    Читает все существовавшие UID из исходного CSV-файла.

    Args:
        csv_path (Path): Путь к исходному CSV-файлу
        config (Dict): Конфигурация приложения (для получения имени колонки UID)
        logger (Optional[logging.Logger]): Логгер для сообщений

    Returns:
        Set[str]: Множество UID, которые уже были в файле
    """
    uids = set()
    uid_header = config["csv"]["headers"]["uid"]

    if logger:
        logger.debug(f"🔍 Поиск UID в колонке: '{uid_header}'")

    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row_num, row in enumerate(reader, start=2):
                uid = row.get(uid_header, "").strip()
                if uid:
                    uids.add(uid)
        if logger:
            logger.debug(
                f"✅ Найдено {len(uids)} существующих UID в {csv_path.name}")
    except Exception as e:
        if logger:
            logger.warning(f"⚠️ Не удалось прочитать оригинальные UID: {e}")
        else:
            print(f"[WARN] Не удалось прочитать оригинальные UID: {e}")
    return uids


def _get_headers(config: Dict) -> tuple:
    """
    Возвращает физические имена колонок из конфигурации.

    Args:
        config (Dict): Конфигурация приложения

    Returns:
        tuple: (path_col, uid_col, cck_col)
    """
    csv_headers = config["csv"]["headers"]
    return (
        csv_headers["path"],
        csv_headers["uid"],
        csv_headers["cck_code"]
    )


def save_modified_xlsx(
    csv_path: Path,
    hierarchy: List[Dict],
    config: Dict,
    logger: Optional[logging.Logger] = None,
    suffix: str = "_modified"
) -> Optional[Path]:
    """
    Сохраняет modified XLSX с цветовой подсветкой старых и новых UID.

    Старые UID (уже были в CSV) — зелёный цвет.
    Новые UID (сгенерированные) — чёрный.

    Args:
        csv_path (Path): Путь к исходному CSV-файлу
        hierarchy (List[Dict]): Список элементов иерархии с полями 'path', 'uid', 'CCK_code'
        config (Dict): Конфигурация приложения
        logger (Optional[logging.Logger]): Логгер для сообщений
        suffix (str): Суффикс для имени файла

    Returns:
        Optional[Path]: Путь к сохранённому XLSX или None при ошибке

    Raises:
        ImportError: Если openpyxl не установлен
    """
    if Workbook is None:
        error_msg = "❌ Для экспорта XLSX требуется openpyxl. Установите: pip install openpyxl"
        if logger:
            logger.error(error_msg)
        else:
            print(error_msg)
        return None

    output_dir = csv_path.parent
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_name = f"{csv_path.stem}{suffix}_{timestamp}.xlsx"
    output_path = output_dir / output_name

    path_col, uid_col, cck_col = _get_headers(config)
    original_uids = _read_original_uids(csv_path, config, logger)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Modified Data"

        # Заголовки
        headers = [path_col, uid_col, cck_col]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Стили
        green_font = Font(color="006400")  # Dark green
        green_fill = PatternFill(start_color="CCFFCC",
                                 end_color="CCFFCC", fill_type="solid")

        for idx, item in enumerate(hierarchy, 2):
            path = item["path"]
            uid = item.get("uid", "").strip()
            ccs_code = item.get("CCK_code", "")

            ws.cell(row=idx, column=1, value=path)
            uid_cell = ws.cell(row=idx, column=2, value=uid)
            ws.cell(row=idx, column=3, value=ccs_code)

            if uid and uid in original_uids:
                uid_cell.font = green_font
                uid_cell.fill = green_fill

        wb.save(output_path)
        if logger:
            logger.info(f"✅ Modified XLSX сохранён: {output_path}")
        return output_path

    except Exception as e:
        if logger:
            logger.error(f"❌ Ошибка при сохранении XLSX: {e}", exc_info=True)
        else:
            print(f"[ERROR] Ошибка при сохранении XLSX: {e}")
        return None


def save_modified_csv(
    csv_path: Path,
    hierarchy: List[Dict],
    config: Dict,
    logger: Optional[logging.Logger] = None,
    suffix: str = "_modified"
) -> Optional[Path]:
    """
    Сохраняет modified CSV с дополнительным столбцом статуса (опционально).

    Args:
        csv_path (Path): Путь к исходному CSV-файлу
        hierarchy (List[Dict]): Список элементов иерархии
        config (Dict): Конфигурация приложения
        logger (Optional[logging.Logger]): Логгер
        suffix (str): Суффикс для имени файла

    Returns:
        Optional[Path]: Путь к сохранённому CSV или None при ошибке
    """
    output_dir = csv_path.parent
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_name = f"{csv_path.stem}{suffix}_{timestamp}.csv"
    output_path = output_dir / output_name

    path_col, uid_col, cck_col = _get_headers(config)
    original_uids = _read_original_uids(csv_path, config, logger)

    fieldnames = [path_col, uid_col, cck_col]
    include_status = config.get("modified_output", {}).get(
        "include_status_column", True)
    if include_status:
        fieldnames.append("status")

    try:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
            writer.writeheader()

            for item in hierarchy:
                path = item["path"]
                uid = item.get("uid", "").strip()
                ccs_code = item.get("CCK_code", "")
                status = "existing" if uid and uid in original_uids else "new"

                row = {
                    path_col: path,
                    uid_col: uid,
                    cck_col: ccs_code
                }
                if include_status:
                    row["status"] = status

                writer.writerow(row)

        if logger:
            logger.info(f"✅ Modified CSV сохранён: {output_path}")
        return output_path

    except Exception as e:
        if logger:
            logger.error(f"❌ Ошибка при сохранении CSV: {e}", exc_info=True)
        else:
            print(f"[ERROR] Ошибка при сохранении CSV: {e}")
        return None


def save_modified_output(
    csv_path: Path,
    hierarchy: List[Dict],
    config: Dict,
    logger: Optional[logging.Logger] = None
) -> Optional[Path]:
    """
    Основная точка входа для экспорта modified-файла.

    Выбирает формат (XLSX или CSV) на основе config.
    Проверяет, включена ли функция.

    Args:
        csv_path (Path): Путь к исходному CSV
        hierarchy (List[Dict]): Данные для экспорта
        config (Dict): Конфигурация
        logger (Optional[logging.Logger]): Логгер

    Returns:
        Optional[Path]: Путь к файлу или None, если отключено/ошибка
    """
    if not config.get("modified_output", {}).get("enabled", True):
        if logger:
            logger.debug("ℹ️ Экспорт modified-файла отключён в config")
        return None

    fmt = config["modified_output"].get("format", "xlsx").lower()
    suffix = config["modified_output"].get("suffix", "_modified")

    if logger:
        logger.info(
            f"📤 Экспорт modified-файла: формат={fmt}, суффикс='{suffix}'")

    if fmt == "csv":
        return save_modified_csv(csv_path, hierarchy, config, logger, suffix)
    else:
        if fmt != "xlsx":
            if logger:
                logger.warning(
                    f"⚠️ Неизвестный формат: {fmt}. Используется xlsx.")
            else:
                print(f"[WARN] Неизвестный формат: {fmt}. Используется xlsx.")
        return save_modified_xlsx(csv_path, hierarchy, config, logger, suffix)
